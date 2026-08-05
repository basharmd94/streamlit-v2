import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from processing import common, purchase
from utils.utils import timed


@st.cache_data(ttl=3600, show_spinner=False)
def _load_tts_final_items(zid: str) -> pd.DataFrame:
    from core.analytics import Analytics
    df = Analytics("final_items_view", zid=zid).data
    return df if df is not None else pd.DataFrame()


def _build_sku_sim_excel(
    calc_df: pd.DataFrame,
    D0: float,
    total_overhead_pool: float,
    days_elapsed: int,
    shipment_name: str = "shipment",
) -> bytes:
    """Return an Excel workbook (bytes) for the SKU Simulator scenario.

    Yellow cells (Avg Price, Days to Clear) are the editable inputs.
    All downstream columns use formulas that recalculate automatically.
    A Totals row at the bottom sums every numeric output column.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SKU Simulator"

    def _f(bold=False, color="000000"):
        return Font(name="Arial", size=10, bold=bold, color=color)

    FILL_HDR    = PatternFill(patternType="solid", fgColor="1F3864")
    FILL_INPUT  = PatternFill(patternType="solid", fgColor="FFFF00")
    FILL_TOTALS = PatternFill(patternType="solid", fgColor="BDD7EE")
    FILL_CONST  = PatternFill(patternType="solid", fgColor="DEEAF1")

    FMT_INT = "#,##0"
    FMT_DEC = "#,##0.00"

    n  = len(calc_df)
    R0 = 6           # first data row
    RE = R0 + n - 1  # last data row
    RT = RE + 1      # totals row

    # ── Constants block (rows 1-3) — D0 lives at B1 ─────────────────────────
    for i, (lbl, val, fmt) in enumerate([
        ("D0 — Daily Overhead Rate", round(D0, 4),            FMT_DEC),
        ("Total Overhead Pool",      round(total_overhead_pool, 2), FMT_DEC),
        ("Days Active (Elapsed)",    int(days_elapsed),        FMT_INT),
    ], start=1):
        ws.cell(row=i, column=1, value=lbl).font = _f(bold=True)
        c = ws.cell(row=i, column=2, value=val)
        c.font = _f(color="000080")
        c.fill = FILL_CONST
        c.number_format = fmt

    # ── Column spec: (header, df_col, is_input, formula_fn, do_sum, fmt) ─────
    def _oh(r):
        return (
            f"=IF(SUM($N${R0}:$N${RE})=0,0,"
            f"$B$1*POWER(0.97,M{r}/60)*M{r}"
            f"*(N{r}/SUM($N${R0}:$N${RE})))"
        )

    COLS = [
        ("Item Code",             "itemcode",            False, None,                       False, "@"),
        ("Item Name",             "itemname",            False, None,                       False, "@"),
        ("Remaining Qty",         "remaining_qty",       False, None,                       False, FMT_INT),
        ("Unit Cost",             "unit_cost",           False, None,                       False, FMT_DEC),
        ("Remaining Cost Value",  "remaining_cost_value",False, None,                       True,  FMT_DEC),
        ("Sold Revenue",          "sold_revenue",        False, None,                       True,  FMT_DEC),
        ("Realized COGS",         "realized_cogs",       False, None,                       True,  FMT_DEC),
        ("Realized GM",           "realized_gm",         False, None,                       True,  FMT_DEC),
        ("Overhead Realized",     "overhead_realized",   False, None,                       True,  FMT_DEC),
        ("Net Profit Realized",   "net_profit_realized", False, None,                       True,  FMT_DEC),
        ("Days Active",           "days_active",         False, None,                       False, FMT_INT),
        ("Avg Price ✏",           "avg_price",           True,  None,                       False, FMT_DEC),
        ("Days to Clear ✏",       "days_to_clear",       True,  None,                       False, FMT_INT),
        ("Proj Remaining Rev",    None,                  False, lambda r: f"=C{r}*L{r}",    True,  FMT_DEC),
        ("Proj Remaining GM",     None,                  False, lambda r: f"=N{r}-E{r}",    True,  FMT_DEC),
        ("Overhead Projected",    None,                  False, _oh,                        True,  FMT_DEC),
        ("Proj Remaining Profit", None,                  False, lambda r: f"=O{r}-P{r}",    True,  FMT_DEC),
        ("Proj Final Profit",     None,                  False, lambda r: f"=J{r}+Q{r}",    True,  FMT_DEC),
    ]

    # ── Header row (row 5) ────────────────────────────────────────────────────
    for ci, (hdr, *_) in enumerate(COLS, start=1):
        c = ws.cell(row=5, column=ci, value=hdr)
        c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill      = FILL_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[5].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(calc_df.iterrows(), start=R0):
        for ci, (_, df_col, is_input, formula_fn, _, fmt) in enumerate(COLS, start=1):
            c = ws.cell(row=ri, column=ci)
            c.number_format = fmt
            if formula_fn:
                c.value = formula_fn(ri)
                c.font  = _f()
            elif is_input:
                try:
                    c.value = float(row[df_col]) if df_col in calc_df.columns else 0.0
                except (ValueError, TypeError):
                    c.value = 0.0
                c.font = _f()
                c.fill = FILL_INPUT
            else:
                raw = row[df_col] if df_col and df_col in calc_df.columns else ""
                if fmt == "@":
                    c.value = str(raw) if raw else ""
                else:
                    try:
                        c.value = float(raw)
                    except (ValueError, TypeError):
                        c.value = ""
                c.font = _f()

    # ── Totals row ────────────────────────────────────────────────────────────
    ws.cell(row=RT, column=1, value="TOTAL").font = _f(bold=True)
    for ci, (_, _, _, _, do_sum, fmt) in enumerate(COLS, start=1):
        c = ws.cell(row=RT, column=ci)
        c.fill = FILL_TOTALS
        c.font = _f(bold=True)
        c.number_format = fmt
        if do_sum:
            cl = get_column_letter(ci)
            c.value = f"=SUM({cl}{R0}:{cl}{RE})"

    # ── Column widths ─────────────────────────────────────────────────────────
    for i, w in enumerate(
        [14, 32, 14, 12, 22, 16, 16, 14, 18, 20, 12, 12, 14, 20, 18, 20, 22, 18],
        start=1,
    ):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Notes ─────────────────────────────────────────────────────────────────
    note_font = Font(name="Arial", size=9, italic=True, color="888888")
    ws.cell(row=RT + 2, column=1,
            value="✏ Yellow cells are editable — change Avg Price or Days to Clear to recalculate.").font = note_font
    ws.cell(row=RT + 3, column=1,
            value=("Overhead = D0 × POWER(0.97, days_to_clear/60) × days_to_clear "
                   "× (SKU proj rev ÷ SUM all proj rev)")).font = note_font

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@st.cache_data(show_spinner=False, ttl=3600)
def _load_inventory_overview(zid_str: str) -> pd.DataFrame:
    from core.analytics import Analytics
    df = Analytics("inventory_overview", zid=zid_str, filters={}).data
    return df if df is not None else pd.DataFrame()


def _render_total_inventory(zid, data_dict):
    """
    Total Inventory Overview — combined 100001 + 100009 stock with
    avg monthly sales velocity and days-to-clear per item.
    """
    st.subheader("Total Inventory Overview")

    with st.spinner("Loading inventory data…"):
        inv_101 = _load_inventory_overview("100001")
        inv_109 = _load_inventory_overview("100009")

    if inv_101.empty and inv_109.empty:
        st.warning("No inventory data available.")
        return

    # ── Resolve packcode for cross-ZID grouping ──────────────────────────
    def _resolve_code(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        d = df.copy()
        d["resolved_code"] = d.apply(
            lambda r: (
                r["packcode"]
                if (
                    r.get("packcode", "")
                    and r["packcode"] not in ("", "NO")
                    and not str(r["packcode"]).upper().startswith("KH")
                )
                else r["item_id"]
            ),
            axis=1,
        )
        return d

    inv_101 = _resolve_code(inv_101)
    inv_109 = _resolve_code(inv_109)

    combined = pd.concat([inv_101, inv_109], ignore_index=True)

    # ── Sum stock by resolved code; carry 100001 metadata as primary ──────
    meta_cols = ["item_id", "item_name", "item_group", "std_price", "min_qty", "min_disc_amt"]
    # 100001 rows preferred for metadata; 100009 rows fill gaps
    meta_101 = inv_101[["resolved_code"] + [c for c in meta_cols if c in inv_101.columns]].drop_duplicates("resolved_code") if not inv_101.empty else pd.DataFrame()
    meta_109 = inv_109[["resolved_code"] + [c for c in meta_cols if c in inv_109.columns]].drop_duplicates("resolved_code") if not inv_109.empty else pd.DataFrame()

    stock_agg = (
        combined.groupby("resolved_code", as_index=False)["stock"]
        .sum()
        .rename(columns={"stock": "total_stock"})
    )

    # Prefer 100001 metadata; fall back to 100009 for items only in 100009
    if not meta_101.empty and not meta_109.empty:
        meta = pd.concat([meta_101, meta_109], ignore_index=True).drop_duplicates("resolved_code", keep="first")
    elif not meta_101.empty:
        meta = meta_101
    else:
        meta = meta_109

    inv_df = stock_agg.merge(meta, on="resolved_code", how="left")

    # ── Avg monthly sales — trailing 12-month window ──────────────────────
    # sales_daily_item is loaded with no date filter (full history), so we
    # cap to the last 12 months ourselves to get a meaningful velocity figure.
    _ROLLING_MONTHS = 12
    sales_raw = data_dict.get("sales_daily_item", pd.DataFrame())
    if not sales_raw.empty and "itemcode" in sales_raw.columns and "quantity" in sales_raw.columns:
        code_map = {}
        if not inv_101.empty:
            for _, r in inv_101[["item_id", "resolved_code"]].drop_duplicates().iterrows():
                code_map[r["item_id"]] = r["resolved_code"]

        s = sales_raw.copy()
        s["date"] = pd.to_datetime(s["date"], errors="coerce")
        s = s.dropna(subset=["date"])
        cutoff = pd.Timestamp.today() - pd.DateOffset(months=_ROLLING_MONTHS)
        s = s[s["date"] >= cutoff]
        s["resolved_code"] = s["itemcode"].map(code_map).fillna(s["itemcode"])

        # Use the actual number of distinct months present (≤12), so a new
        # product that only appeared 3 months ago isn't penalised by dividing by 12.
        num_months = max(s["date"].dt.to_period("M").nunique(), 1)
        sales_agg = (
            s.groupby("resolved_code", as_index=False)["quantity"]
            .sum()
            .assign(avg_monthly_sales=lambda d: d["quantity"] / num_months)
            [["resolved_code", "avg_monthly_sales"]]
        )
        inv_df = inv_df.merge(sales_agg, on="resolved_code", how="left")
    else:
        inv_df["avg_monthly_sales"] = 0.0
        num_months = 0

    inv_df["avg_monthly_sales"] = inv_df["avg_monthly_sales"].fillna(0.0)

    # ── Days to clear ──────────────────────────────────────────────────────
    inv_df["days_to_clear"] = inv_df.apply(
        lambda r: round(r["total_stock"] / r["avg_monthly_sales"] * 30, 1)
        if r["avg_monthly_sales"] > 0 else None,
        axis=1,
    )

    # ── Display ────────────────────────────────────────────────────────────
    col_map = {
        "item_id":           "Item Code",
        "item_name":         "Item Name",
        "item_group":        "Item Group",
        "total_stock":       "Stock",
        "avg_monthly_sales": "Avg Monthly Sales",
        "days_to_clear":     "Days to Clear",
        "std_price":         "Std Price",
        "min_disc_amt":      "Min Disc Amt",
        "min_qty":           "Min Qty",
    }
    display_df = (
        inv_df
        .rename(columns=col_map)
        [[c for c in col_map.values() if c in inv_df.rename(columns=col_map).columns]]
        .reset_index(drop=True)
    )

    # Search filter
    search = st.text_input("Search item name or group", "")
    if search:
        mask = (
            display_df.get("Item Name", pd.Series(dtype=str)).str.contains(search, case=False, na=False)
            | display_df.get("Item Group", pd.Series(dtype=str)).str.contains(search, case=False, na=False)
            | display_df.get("Item Code", pd.Series(dtype=str)).str.contains(search, case=False, na=False)
        )
        display_df = display_df[mask]

    st.caption(
        f"Stock = 100001 + 100009 combined (cross-ZID packcode merge). "
        f"Avg Monthly Sales = trailing 12-month window ({num_months} month(s) with data). "
        f"Days to Clear = Stock ÷ Avg Monthly Sales × 30."
    )

    fmt = {c: "{:,.0f}" for c in ["Stock", "Avg Monthly Sales", "Std Price", "Min Disc Amt", "Min Qty"]}
    fmt["Days to Clear"] = "{:,.1f}"
    existing_fmt = {k: v for k, v in fmt.items() if k in display_df.columns}

    st.dataframe(
        display_df.style.format(existing_fmt, na_rep="—"),
        use_container_width=True,
        hide_index=True,
    )

    st.download_button(
        label="📥 Download as CSV",
        data=display_df.to_csv(index=False).encode("utf-8"),
        file_name="total_inventory_overview.csv",
        mime="text/csv",
    )


_TTS_COL_LABELS = {
    "itemcode": "Item Code",
    "itemname": "Item Name",
    "itemgroup": "Group",
    "incoming_qty": "Incoming Qty",
    "stock": "Stock",
    "P50": "P50 (days)",
    "P75": "P75 (days)",
    "P90": "P90 (days)",
    "P95": "P95 (days)",
    "n": "Data Points (n)",
    "flag": "Flag",
}

_TTS_FLAG_STYLE = {
    "Dead stock": "background-color: #ffd6d6; color: #8b0000;",
    "Low confidence": "background-color: #fff3cd; color: #856404;",
    "No data": "background-color: #e2e3e5; color: #383d41;",
}


def _render_time_to_sell(zid: str, data_dict: dict) -> None:
    """Time to Sell percentile table — upcoming open shipments or current inventory."""
    from processing.next_month_target import get_open_shipments, get_shipment_items
    from processing.purchase_batch import build_time_to_sell_percentiles

    purchase_df = data_dict.get("purchase_batches", pd.DataFrame())
    sales_df = data_dict.get("sales_daily_item", pd.DataFrame())

    if purchase_df is None or purchase_df.empty:
        st.warning("No purchase data loaded.")
        return

    with st.spinner("Computing sell-through percentiles…"):
        pct_df = build_time_to_sell_percentiles(purchase_df, sales_df)

    view = st.radio(
        "Analysis scope",
        ["📦 Upcoming Shipment", "🏭 Current Inventory"],
        horizontal=True,
        key="tts_inner_radio",
    )

    st.divider()

    def _show_table(df: pd.DataFrame, qty_col: str | None = None) -> None:
        id_cols = ["itemcode", "itemname", "itemgroup"]
        pct_cols = ["P50", "P75", "P90", "P95", "n", "flag"]
        ordered = id_cols + ([qty_col] if qty_col and qty_col in df.columns else []) + pct_cols
        display_df = df[[c for c in ordered if c in df.columns]].rename(columns=_TTS_COL_LABELS)
        if display_df.empty:
            st.info("No data to display.")
            return

        def _flag_style(val):
            return _TTS_FLAG_STYLE.get(val, "")

        try:
            styled = display_df.style.applymap(_flag_style, subset=["Flag"])
        except Exception:
            styled = display_df.style

        st.dataframe(styled, use_container_width=True, hide_index=True)
        st.download_button(
            "📥 Download CSV",
            display_df.to_csv(index=False).encode(),
            file_name="time_to_sell.csv",
            mime="text/csv",
            key="tts_download",
        )

    # ---- Upcoming Shipment ----
    if view == "📦 Upcoming Shipment":
        open_df = get_open_shipments(purchase_df)
        if open_df.empty:
            st.info("No open shipments found.")
            return

        open_df["label"] = (
            open_df["shipmentname"].astype(str)
            + "  ("
            + open_df["n_items"].astype(str) + " items, "
            + open_df["total_qty"].apply(lambda x: f"{x:,.0f}") + " units)"
        )

        selected_labels = st.multiselect(
            "Select shipment(s)",
            options=open_df["label"].tolist(),
            default=[open_df["label"].iloc[0]],
            key="tts_shipment_select",
        )
        if not selected_labels:
            st.info("Select at least one shipment.")
            return

        sel_rows = open_df[open_df["label"].isin(selected_labels)]
        selections = list(zip(sel_rows["zid"], sel_rows["shipmentname"]))
        items_df = get_shipment_items(purchase_df, selections)

        if items_df.empty:
            st.warning("No items found for the selected shipment(s).")
            return

        items_df["itemcode"] = items_df["itemcode"].astype(str).str.strip()
        merged = items_df.merge(pct_df, on="itemcode", how="left", suffixes=("", "_pct"))
        for dup in [c for c in merged.columns if c.endswith("_pct")]:
            merged = merged.drop(columns=[dup])
        merged["itemgroup"] = merged.get("itemgroup", pd.Series([""] * len(merged))).fillna("")
        merged["flag"] = merged["flag"].fillna("No data")
        merged["n"] = pd.to_numeric(merged["n"], errors="coerce").fillna(0).astype(int)

        st.caption(f"{len(merged)} item(s) in selected shipment(s)")
        _show_table(merged, qty_col="incoming_qty")

    # ---- Current Inventory ----
    else:
        inv_df = _load_tts_final_items(str(zid))
        if inv_df is None or inv_df.empty:
            st.warning("No current inventory data.")
            return

        if "item_id" in inv_df.columns:
            inv_df = inv_df.rename(columns={"item_id": "itemcode", "item_name": "itemname", "item_group": "itemgroup"})

        inv_df["itemcode"] = inv_df["itemcode"].astype(str).str.strip()
        inv_df = inv_df[pd.to_numeric(inv_df.get("stock", 0), errors="coerce").fillna(0) > 0].copy()

        merged = inv_df.merge(pct_df, on="itemcode", how="left", suffixes=("", "_pct"))
        for col in ["itemname", "itemgroup"]:
            pct_col = col + "_pct"
            if pct_col in merged.columns:
                merged[col] = merged[col].where(merged[col].notna() & (merged[col] != ""), merged[pct_col])
                merged = merged.drop(columns=[pct_col])
        merged["flag"] = merged["flag"].fillna("No data")
        merged["n"] = pd.to_numeric(merged["n"], errors="coerce").fillna(0).astype(int)

        st.caption(f"{len(merged)} item(s) currently in stock for ZID {zid}")
        _show_table(merged, qty_col="stock")


@timed
def display_purchase_analysis_page(current_page, zid, data_dict):

    mode = st.radio(
        "Purchase View",
        [
            "Purchase Cohort & Requisition",
            "Batch Profitability & Capital Engine",
            "Total Inventory Overview",
            "Time to Sell Analysis",
        ],
        horizontal=True,
        index=0,
    )

    # -----------------------------
    # MODE 3: Total Inventory Overview
    # -----------------------------
    if mode == "Total Inventory Overview":
        _render_total_inventory(zid, data_dict)
        return

    # -----------------------------
    # MODE 4: Time to Sell Analysis
    # -----------------------------
    if mode == "Time to Sell Analysis":
        _render_time_to_sell(str(zid), data_dict)
        return

    # -----------------------------
    # MODE 1: Existing cohort logic (unchanged)
    # -----------------------------
    if mode == "Purchase Cohort & Requisition":
        options = [i for i in range(10)]
        default_option = 2
        default_index = options.index(default_option)
        selected_time = st.selectbox("Select Time Frame", options, index=default_index)

        sales_df, purchase_df, year_ago = common.time_filtered_data_purchase(
            data_dict['sales_daily_item'], data_dict['purchase_batches'], selected_time
        )
        cohort_df = purchase.main_purchase_product_cohort_process(sales_df, purchase_df)

        selected_products = st.multiselect("Select Product", common.update_pair_options(cohort_df,'itemcode','itemname'))
        if selected_products:
            selected_itemcodes = [x.split(" - ")[0] for x in selected_products]
            cohort_df = cohort_df[cohort_df['itemcode'].isin(selected_itemcodes)]

        cohort_df = cohort_df.applymap(common.handle_infinity_and_round).fillna(0)
        st.markdown("Product-Based Purchase Cohort")
        st.write(cohort_df, use_container_width=True)
        st.write(common.create_download_link(cohort_df, "purchase_cohort.xlsx"), unsafe_allow_html=True)

        if st.button("Generate Purchase Requirement"):
            result_df = purchase.generate_cohort(
                data_dict['purchase_batches'],
                year_ago,
                data_dict['stock_movement'],
                sales_df,
                cohort_df
            )
            st.markdown("Generated Purchase Requisition")
            st.write(result_df, use_container_width=True)
            st.write(common.create_download_link(result_df, "purchase_requisition.xlsx"), unsafe_allow_html=True)

        return

    else:
        # -----------------------------
        # MODE 2: Batch Profitability & Capital Engine
        # -----------------------------
        st.subheader("Batch Profitability & Capital Engine")

        purchase_df = data_dict.get("purchase_batches", pd.DataFrame())
        if purchase_df is None or purchase_df.empty:
            st.warning("No purchase data loaded.")
            return

             # ============================================================
        # SECTION RADIO (ONLY ONE SECTION COMPUTES PER RERUN)
        # Make Accounts Explorer the default landing section
        # ============================================================
        engine_section = st.radio(
            "Engine Section",
            [
                "Accounts Explorer (Overhead)",
                "Inventory Check",
                "Warehouse Snapshot",
                "Batch Profitability",
                "SKU Simulator",
            ],
            horizontal=True,
            index=0,
            key="purchase_engine_section",
        )

        # --------- Shipment selector ----------
        ship_df = purchase_df[["zid", "shipmentname", "povoucher", "grnvoucher", "combinedate"]].copy()
        ship_df["shipmentname"] = ship_df["shipmentname"].astype(str).fillna("").str.strip()
        ship_df = ship_df[ship_df["shipmentname"] != ""].copy()

        shipment_options = (
            ship_df[["shipmentname"]]
            .drop_duplicates()
            .sort_values("shipmentname")["shipmentname"]
            .tolist()
        )

        if not shipment_options:
            st.warning("No shipmentname found in purchase data.")
            return

        selected_shipment = st.selectbox(
            "Select Shipment (bridges 100001 + 100009)",
            shipment_options,
            index=0,
            key="purchase_selected_shipment",
            disabled=(engine_section != "Accounts Explorer (Overhead)"),
        )

        # Combinedate resolve
        _sel_rows = ship_df[ship_df["shipmentname"] == selected_shipment].copy()
        selected_combinedate = pd.to_datetime(_sel_rows["combinedate"], errors="coerce").min()

        if pd.isna(selected_combinedate):
            st.error("Could not resolve combinedate.")
            return

        selected_combinedate = selected_combinedate.normalize()

        # Display IP/GRN transparency
        ship_pick = ship_df[ship_df["shipmentname"] == selected_shipment].copy()

        info_cols = st.columns(2)
        with info_cols[0]:
            st.caption("100001 (Trading)")
            st.write({
                "IP": ship_pick[ship_pick["zid"].astype(str) == "100001"]["povoucher"].dropna().unique().tolist(),
                "GRN": ship_pick[ship_pick["zid"].astype(str) == "100001"]["grnvoucher"].dropna().unique().tolist()
            })

        with info_cols[1]:
            st.caption("100009 (Packaging)")
            st.write({
                "IP": ship_pick[ship_pick["zid"].astype(str) == "100009"]["povoucher"].dropna().unique().tolist(),
                "GRN": ship_pick[ship_pick["zid"].astype(str) == "100009"]["grnvoucher"].dropna().unique().tolist()
            })

        st.divider()

        # ---------------------------------------------------
        # Warehouse Selection (Shared across all sections)
        # ---------------------------------------------------
        wh_opts = purchase.get_all_warehouse_options(data_dict["stock_movement"])

        sel_wh_100001 = st.multiselect(
            "Warehouses (100001)",
            options=wh_opts.get("100001", []),
            default=wh_opts.get("100001", []),
        )

        sel_wh_100009 = st.multiselect(
            "Warehouses (100009)",
            options=wh_opts.get("100009", []),
            default=wh_opts.get("100009", []),
        )

        override_wh = {
            "100001": sel_wh_100001,
            "100009": sel_wh_100009,
        }
        # ============================================================
        # 1 INVENTORY CHECK
        # ============================================================

        if engine_section == "Inventory Check":

            tables = purchase.build_shipment_inventory_tables(
                purchase_df=data_dict["purchase_batches"],
                stock_movement_df=data_dict["stock_movement"],
                sales_df=data_dict["sales_daily_item"],
                returns_df=data_dict["returns_daily_item"],
                shipmentname=selected_shipment,
                project=st.session_state.proj,
                zid_deplete="100001",
            )

            # -----------------------------
            # Save audit results for reuse in Batch Profitability
            # -----------------------------
            st.session_state["invcheck_tables"] = tables
            st.session_state["invcheck_reconcile"] = tables.get("reconcile_sales_vs_stock", pd.DataFrame())

            with st.expander("Inventory Check Tables", expanded=True):
                st.subheader("Arrival Check — 100001")
                st.dataframe(tables["arrival_check_100001_only"], use_container_width=True)

                st.subheader("Arrival Check — 100009 Items")
                st.dataframe(tables["arrival_check_100009_items"], use_container_width=True)

                st.subheader("Sales vs Stock Reconciliation")
                st.dataframe(tables["reconcile_sales_vs_stock"], use_container_width=True)

                # st.subheader("Warehouse Breakdown")
                # st.dataframe(tables["warehouse_breakdown"], use_container_width=True)

            return


        # ============================================================
        # 2 ACCOUNTS EXPLORER
        # ============================================================

        if engine_section == "Accounts Explorer (Overhead)":

            opts = purchase.build_accounts_selector_options(
                glmst_df=data_dict["glmst_simple"],
                hierarchy_path="data/hierarchy.json",
            )

            level_choice = st.radio(
                "Selection level",
                ["Level 0", "Level 1"],
                horizontal=True,
                index=0,
            )

            level_map = {"Level 0": 0, "Level 1": 1}
            level = level_map[level_choice]

            if level_choice == "Level 0":
                selections = st.multiselect(
                    "Select ac_code(s)",
                    opts["level0_options"],
                    default=opts["level0_options"],
                )
                selections = [s.split(" ", 1)[0].strip() for s in selections]

            else:
                selections = st.multiselect(
                    "Select Level 1 head(s)",
                    opts["level1_options"],
                    default=opts["level1_options"],
                )

            # 08020003 always included — negative GL value reduces the overhead pool
            revenue_selections = ("08020003",)
            st.caption("Revenue adjustment applied: `08020003` (intercompany services — negative value reduces overhead)")

            show_details = st.checkbox("Show daily ratio diagnostics")

            # Split gl_overhead_daily by ZID so each entity gets its own table
            _gl_all = data_dict.get("gl_overhead_daily", pd.DataFrame())
            def _gl_for_zid(df, zid_val):
                if df is None or df.empty:
                    return pd.DataFrame()
                d = df.copy()
                d["zid"] = d["zid"].astype(str).str.strip()
                return d[d["zid"] == str(zid_val)].reset_index(drop=True)

            _common_kw = dict(
                purchase_df=data_dict["purchase_batches"],
                stock_movement_df=data_dict["stock_movement"],
                glmst_df=data_dict["glmst_simple"],
                hierarchy_path="data/hierarchy.json",
                shipmentname=selected_shipment,
                level=level,
                selections=tuple(selections),
                include_details=show_details,
                zids_inventory=["100001", "100009"],
                warehouse_filters=override_wh,
                warehouse_json_path="data/warehouse_filters.json",
                revenue_selections=revenue_selections,
            )

            out_100001 = purchase.build_accounts_overhead_summary(
                gl_overhead_df=_gl_for_zid(_gl_all, "100001"), **_common_kw
            )
            out_100009 = purchase.build_accounts_overhead_summary(
                gl_overhead_df=_gl_for_zid(_gl_all, "100009"), **_common_kw
            )

            alloc_100001 = float(out_100001["totals"].get("overhead_for_shipment_sum", 0.0))
            alloc_100009 = float(out_100009["totals"].get("overhead_for_shipment_sum", 0.0))
            pool_100001  = float(out_100001["totals"].get("overhead_total_sum", 0.0))
            pool_100009  = float(out_100009["totals"].get("overhead_total_sum", 0.0))
            db_overhead  = alloc_100001 + alloc_100009

            # ── Per-entity tables ──────────────────────────────────────────────
            st.markdown("#### 🏢 HMBR Tools & Chemicals — 100001")
            if not out_100001["summary_df"].empty:
                st.dataframe(out_100001["summary_df"], use_container_width=True)
            else:
                st.info("No overhead data for 100001 in this shipment window.")

            if show_details and out_100001["details_df"] is not None and not out_100001["details_df"].empty:
                with st.expander("Daily Diagnostics — 100001", expanded=False):
                    st.dataframe(out_100001["details_df"], use_container_width=True)

            st.markdown("#### 📦 Gulshan Packaging — 100009")
            if not out_100009["summary_df"].empty:
                st.dataframe(out_100009["summary_df"], use_container_width=True)
            else:
                st.info("No overhead data for 100009 in this shipment window.")

            if show_details and out_100009["details_df"] is not None and not out_100009["details_df"].empty:
                with st.expander("Daily Diagnostics — 100009", expanded=False):
                    st.dataframe(out_100009["details_df"], use_container_width=True)

            # ── Totals breakdown ───────────────────────────────────────────────
            st.markdown("#### 📊 Overhead Totals")
            _tc1, _tc2, _tc3 = st.columns(3)
            with _tc1:
                st.metric("100001 — Total Pool", f"{pool_100001:,.0f}")
                st.metric("100001 — Shipment Allocated", f"{alloc_100001:,.0f}")
            with _tc2:
                st.metric("100009 — Total Pool", f"{pool_100009:,.0f}")
                st.metric("100009 — Shipment Allocated", f"{alloc_100009:,.0f}")
            with _tc3:
                st.metric("Combined — Total Pool", f"{pool_100001 + pool_100009:,.0f}")
                st.metric("Combined — Shipment Allocated", f"{db_overhead:,.0f}")

            st.markdown("### Overhead Add-ons (optional)")

            c1, c2 = st.columns(2)
            with c1:
                use_vat = st.checkbox("Add VAT overhead (%) on sales", value=st.session_state.get("use_vat_overhead", False))
                vat_pct = st.number_input("VAT %", min_value=0.0, max_value=50.0, value=float(st.session_state.get("vat_pct", 0.0)), step=0.5)
            with c2:
                use_manual = st.checkbox("Add manual overhead (BDT)", value=st.session_state.get("use_manual_overhead", False))
                manual_overhead_value = st.number_input("Manual overhead value", min_value=0.0, value=float(st.session_state.get("manual_overhead_value", 0.0)), step=100.0)

            st.session_state["use_vat_overhead"] = use_vat
            st.session_state["vat_pct"] = vat_pct if use_vat else 0.0
            st.session_state["use_manual_overhead"] = use_manual
            st.session_state["manual_overhead_value"] = manual_overhead_value if use_manual else 0.0

            # We show VAT estimate using last computed revenue if available (exact VAT will be applied in engine)
            last_rev = float(st.session_state.get("last_shipment_realized_revenue", 0.0))
            vat_est = (st.session_state["vat_pct"] / 100.0) * max(0.0, last_rev)
            manual_val = float(st.session_state.get("manual_overhead_value", 0.0))

            st.write({
                "DB overhead (allocated)": round(db_overhead, 2),
                "VAT overhead (estimate; exact in engine)": round(vat_est, 2),
                "Manual overhead": round(manual_val, 2),
                "Total overhead passed to engine (estimate)": round(db_overhead + vat_est + manual_val, 2),
            })

            st.session_state["shipment_overhead_total"] = db_overhead

            with st.expander("📐 How Accounts Explorer calculates overhead", expanded=False):
                st.markdown("""
**Purpose:** Allocates GL overhead expenses (05/06/07 accounts) to this shipment proportionally, based on the shipment's share of total inventory value each day it is active.

**Date range:** From the shipment's `combinedate` to either the date the shipment is fully depleted (stock-depletion model) or today if still open.

---
**Daily allocation formula:**

| Variable | Meaning |
|---|---|
| `shipment_value_cost` | Cost of shipment units still on-hand on that day (depletion model) |
| `total_inventory_value_cost` | Cumulative on-hand cost across all items in selected warehouses |
| `ratio` | `shipment_value_cost ÷ total_inventory_value_cost` |
| `overhead_total_for_day` | Sum of selected GL expense postings on that day |
| `overhead_allocated_for_day` | `overhead_total_for_day × ratio` |

**Total allocated to shipment** = Σ `overhead_allocated_for_day` over all active days → stored as `shipment_overhead_total` and passed to Batch Profitability.

---
**Revenue Adjustments:** When a revenue account is selected here, its GL value (negative, since it is an income credit) is added to the overhead pool for that day. This reduces `overhead_total_for_day` by the revenue amount, which in turn lowers the allocated overhead for the shipment.

*Example: Revenue adjustment posting = −10,000 on a day with 50,000 in expenses → net overhead_for_day = 40,000 → allocated at 10% ratio = 4,000 instead of 5,000.*

---
**Numerical example (simplified):**

| Day | Expenses (GL) | Rev Adj | Net OH | Ship Value | Total Inv | Ratio | Allocated |
|---|---|---|---|---|---|---|---|
| 1 | 50,000 | −10,000 | 40,000 | 10,00,000 | 1,00,00,000 | 10.0% | 4,000 |
| 2 | 30,000 | 0 | 30,000 | 9,80,000 | 1,00,00,000 | 9.8% | 2,940 |
| 3 | 45,000 | 0 | 45,000 | 9,50,000 | 1,00,00,000 | 9.5% | 4,275 |

**Total allocated → 11,215** → this is `shipment_overhead_total` fed into Batch Profitability.
""")

            return

        # ============================================================
        # 3 BATCH PROFITABILITY
        # ============================================================

        if engine_section == "Batch Profitability":

            if "shipment_overhead_total" not in st.session_state:
                st.warning("Please run Accounts Explorer first to compute shipment overhead.")
                return

            shipment_overhead_total = float(st.session_state.get("shipment_overhead_total", 0.0))
            vat_pct = float(st.session_state.get("vat_pct", 0.0))
            manual_overhead_value = float(st.session_state.get("manual_overhead_value", 0.0))

            result_df = purchase.run_batch_profitability_engine(
                purchase_df=data_dict["purchase_batches"],
                sales_df=data_dict["sales_daily_item"],
                returns_df=data_dict["returns_daily_item"],
                stock_movement_df=data_dict["stock_movement"],
                hierarchy_path="data/hierarchy.json",
                shipmentname=selected_shipment,
                discount_pct=0.0,
                zid_deplete="100001",
                shipment_overhead_total=shipment_overhead_total,
                vat_pct=vat_pct,
                manual_overhead_value=manual_overhead_value,
                inventory_tables=st.session_state.get("invcheck_tables"),
            )

            st.session_state["last_batch_df"] = result_df.copy()
            st.session_state["last_shipment_realized_revenue"] = float(result_df["sold_revenue"].sum()) if not result_df.empty else 0.0

            st.dataframe(result_df, use_container_width=True)

            if result_df is not None and not result_df.empty:
                sum_cols = [
                    "sold_revenue", "realized_cogs", "realized_gm",
                    "overhead_realized", "net_profit_realized",
                    "remaining_cost_value", "proj_remaining_revenue", "proj_remaining_gm",
                    "overhead_projected", "Proj_remaining_profit", "proj_final_profit",
                ]
                totals = {c: float(result_df[c].sum()) for c in sum_cols if c in result_df.columns}

                # averages (simple + weighted)
                vel = result_df["velocity"].replace([np.inf, -np.inf], np.nan)
                dtc = result_df["days_to_clear"].replace([np.inf, -np.inf], np.nan)

                simple_avg_velocity = float(vel.mean(skipna=True))
                simple_avg_days_to_clear = float(dtc.mean(skipna=True))

                # Weighted: velocity weighted by sold_qty; days_to_clear weighted by remaining_qty
                sold_w = result_df["sold_qty"].clip(lower=0.0)
                rem_w = result_df["remaining_qty"].clip(lower=0.0)

                w_avg_velocity = float((vel.fillna(0.0) * sold_w).sum() / sold_w.sum()) if sold_w.sum() > 0 else 0.0
                w_avg_days_to_clear = float((dtc.fillna(0.0) * rem_w).sum() / rem_w.sum()) if rem_w.sum() > 0 else 0.0

                st.markdown("### Shipment Totals & Averages")
                st.write({k: round(v, 2) for k, v in totals.items()})

                st.write({
                    "Avg velocity (simple)": round(simple_avg_velocity, 4),
                    "Avg days_to_clear (simple)": round(simple_avg_days_to_clear, 2),
                    "Avg velocity (weighted by sold_qty)": round(w_avg_velocity, 4),
                    "Avg days_to_clear (weighted by remaining_qty)": round(w_avg_days_to_clear, 2),
                    "Shipment Overhead Total From Account Explorer": shipment_overhead_total,
                })
                with st.expander("📐 How Batch Profitability calculates profit", expanded=False):
                    st.markdown("""
**FIFO allocation rule:** Sales are allocated to batches of the same SKU in chronological order. Earlier batches get consumed first. A later batch can only absorb sales that exceed the remaining stock of all earlier batches.

---
**Per-batch fields:**

| Field | Formula |
|---|---|
| `sold_qty` | Units of this batch consumed by sales (FIFO, net of returns) |
| `sold_revenue` | `sold_qty × weighted avg selling price` during the batch's active period |
| `realized_cogs` | `sold_qty × unit_cost` |
| `realized_gm` | `sold_revenue − realized_cogs` |

**Overhead (realized portion):**

| Field | Formula |
|---|---|
| `D0` | `total_overhead_pool ÷ max(days_active)` across all SKUs in this shipment |
| `realized_share` | `SKU sold_revenue ÷ shipment total sold_revenue` (falls back to cost share if 0) |
| `overhead_realized` | `D0 × days_active × realized_share` |
| `net_profit_realized` | `realized_gm − overhead_realized` |

**Projected (remaining stock):**

| Field | Formula |
|---|---|
| `proj_remaining_revenue` | `remaining_qty × avg_price` (avg selling price of this SKU) |
| `proj_remaining_gm` | `proj_remaining_revenue − (remaining_qty × unit_cost)` |
| `remaining_share` | `SKU proj_remaining_revenue ÷ total proj remaining revenue` |
| `overhead_projected` | `D0 × 0.97^(days_to_clear ÷ 60) × days_to_clear × remaining_share` |
| `Proj_remaining_profit` | `proj_remaining_gm − overhead_projected` |
| `proj_final_profit` | `net_profit_realized + Proj_remaining_profit` |

*The 0.97 decay factor reflects that overhead pressure on remaining stock diminishes as a shipment ages — the cost base is already partially recouped.*

---
**Numerical example:**

- Batch: 100 units purchased at BDT 500/unit → `batch_cost = 50,000`
- 70 units sold at avg BDT 750 → `sold_revenue = 52,500`; `realized_cogs = 35,000`; `realized_gm = 17,500`
- `D0 = 100/day`; `days_active = 90`; `realized_share = 40%`
  → `overhead_realized = 100 × 90 × 0.40 = 3,600`
  → `net_profit_realized = 17,500 − 3,600 = 13,900`
- Remaining: 30 units; `days_to_clear = 45`; `remaining_share = 25%`
  → `proj_remaining_revenue = 30 × 750 = 22,500`; `proj_remaining_gm = 22,500 − 15,000 = 7,500`
  → `overhead_projected = 100 × 0.97^(45/60) × 45 × 0.25 ≈ 1,084`
  → `Proj_remaining_profit = 7,500 − 1,084 = 6,416`
  → `proj_final_profit = 13,900 + 6,416 = 20,316`
""")
            return
        # ============================================================
        # 4 WAREHOUSE SNAPSHOT
        # ============================================================

        if engine_section == "Warehouse Snapshot":

            warehouse_basis = st.radio(
                "Warehouse valuation date",
                ["Shipment arrival (combinedate)", "Today"],
                horizontal=True,
            )

            as_of_dt = selected_combinedate if warehouse_basis == "Shipment arrival (combinedate)" else pd.Timestamp.today().normalize()

            wh_value_df = purchase.build_warehouse_total_value_table(
                stock_movement_df=data_dict["stock_movement"],
                as_of_date=as_of_dt,
                zids=["100001", "100009"],
                warehouse_filters=override_wh,
                warehouse_json_path="data/warehouse_filters.json",
            )

            st.dataframe(wh_value_df, use_container_width=True)

            st.metric("Total Inventory Value",
                    round(wh_value_df["totalvalue"].sum(), 2))

            return

        # ============================================================
        # 3 SKU Simulator
        # ============================================================

        if engine_section == "SKU Simulator":
            st.subheader("Scenario Worksheet")

            base_df = st.session_state.get("last_batch_df")

            if base_df is None or base_df.empty:
                st.warning("Run Batch Profitability first.")
                return

            # ---------------------------------------------
            # Base worksheet shown to user
            # ---------------------------------------------
            work_df = base_df.copy()
            work_df["avg_price"] = pd.to_numeric(work_df["avg_price"], errors="coerce").fillna(0.0)
            work_df["days_to_clear"] = pd.to_numeric(work_df["days_to_clear"], errors="coerce").fillna(0.0)

            editable_df = st.data_editor(
                work_df,
                use_container_width=True,
                num_rows="fixed",
                hide_index=True,
                column_config={
                    "avg_price": st.column_config.NumberColumn("avg_price", step=1.0, format="%.2f"),
                    "days_to_clear": st.column_config.NumberColumn("days_to_clear", step=1.0, format="%.2f"),
                },
                disabled=[c for c in work_df.columns if c not in ["avg_price", "days_to_clear"]],
                key="scenario_worksheet_editor",
            )

            # ---------------------------------------------
            # Recompute scenario using edited values
            # BUT anchor all share/pool logic to original batch result
            # ---------------------------------------------
            calc_df = base_df.copy()
            share_base_df = base_df.copy()

            calc_df["avg_price"] = pd.to_numeric(editable_df["avg_price"], errors="coerce").fillna(0.0)
            calc_df["days_to_clear"] = (
                pd.to_numeric(editable_df["days_to_clear"], errors="coerce")
                .fillna(0.0)
                .clip(lower=0.0, upper=730.0)
            )

            # scenario price follows edited avg price
            calc_df["scenario_price"] = calc_df["avg_price"]

            calc_df["remaining_qty"] = pd.to_numeric(calc_df["remaining_qty"], errors="coerce").fillna(0.0)
            calc_df["remaining_cost_value"] = pd.to_numeric(calc_df["remaining_cost_value"], errors="coerce").fillna(0.0)
            calc_df["net_profit_realized"] = pd.to_numeric(calc_df["net_profit_realized"], errors="coerce").fillna(0.0)
            calc_df["days_active"] = pd.to_numeric(calc_df["days_active"], errors="coerce").fillna(1.0)

            calc_df["proj_remaining_revenue"] = calc_df["remaining_qty"] * calc_df["scenario_price"]
            calc_df["proj_remaining_gm"] = calc_df["proj_remaining_revenue"] - calc_df["remaining_cost_value"]

            # ---------------------------------------------
            # Overhead base must match original Batch Profitability engine
            # ---------------------------------------------
            shipment_overhead_total = float(st.session_state.get("shipment_overhead_total", 0.0))
            vat_pct = float(st.session_state.get("vat_pct", 0.0))
            manual_overhead_value = float(st.session_state.get("manual_overhead_value", 0.0))

            # VAT overhead in the original engine is based on realized sold revenue
            total_sold_revenue_base = float(
                pd.to_numeric(base_df["sold_revenue"], errors="coerce").fillna(0.0).sum()
            )

            vat_overhead_value = (vat_pct / 100.0) * max(0.0, total_sold_revenue_base)

            total_overhead_pool = (
                float(shipment_overhead_total)
                + float(vat_overhead_value)
                + float(manual_overhead_value)
            )

            days_elapsed = int(
                pd.to_numeric(base_df["days_active"], errors="coerce").fillna(1).max()
            )
            days_elapsed = max(1, days_elapsed)

            D0 = total_overhead_pool / float(days_elapsed)

            # ---------------------------------------------
            # Remaining-share basis anchored to original sheet shape
            # but with edited projected revenue values from current worksheet
            # ---------------------------------------------
            total_proj_remaining_revenue = float(
                pd.to_numeric(calc_df["proj_remaining_revenue"], errors="coerce").fillna(0.0).sum()
            )

            if total_proj_remaining_revenue > 0:
                share_rem = (
                    pd.to_numeric(calc_df["proj_remaining_revenue"], errors="coerce").fillna(0.0)
                    / total_proj_remaining_revenue
                )
            else:
                denom = float(
                    pd.to_numeric(share_base_df["remaining_cost_value"], errors="coerce").fillna(0.0).sum()
                )
                if denom > 0:
                    share_rem = (
                        pd.to_numeric(calc_df["remaining_cost_value"], errors="coerce").fillna(0.0)
                        / denom
                    )
                else:
                    share_rem = 0.0

            # ---------------------------------------------
            # Same projected overhead formula as batch engine
            # overhead_projected = D0 * (0.97)^(days_to_clear/60) * days_to_clear * remaining_share
            # ---------------------------------------------
            dclear = (
                pd.to_numeric(calc_df["days_to_clear"], errors="coerce")
                .fillna(0.0)
                .clip(lower=0.0, upper=730.0)
            )

            decay_factor = np.power(0.97, dclear / 60.0)

            calc_df["overhead_projected"] = D0 * decay_factor * dclear * share_rem
            calc_df["Proj_remaining_profit"] = calc_df["proj_remaining_gm"] - calc_df["overhead_projected"]
            calc_df["proj_final_profit"] = calc_df["net_profit_realized"] + calc_df["Proj_remaining_profit"]

            # ---------------------------------------------
            # Totals row
            # Only sum revenue / cost / overhead / profit fields
            # ---------------------------------------------
            total_cols = [
                "sold_revenue",
                "realized_cogs",
                "realized_gm",
                "overhead_realized",
                "net_profit_realized",
                "remaining_cost_value",
                "proj_remaining_revenue",
                "proj_remaining_gm",
                "overhead_projected",
                "Proj_remaining_profit",
                "proj_final_profit",
            ]

            totals_row = {c: "" for c in calc_df.columns}
            totals_row["shipmentname"] = "TOTAL"
            totals_row["batch_id"] = ""
            totals_row["itemcode"] = ""
            totals_row["itemname"] = ""
            totals_row["combinedate"] = ""
            totals_row["batch_end_date"] = ""
            totals_row["is_closed"] = ""
            totals_row["initial_qty"] = ""
            totals_row["sold_qty"] = ""
            totals_row["remaining_qty"] = ""
            totals_row["threshold_qty"] = ""
            totals_row["unit_cost"] = ""
            totals_row["avg_price"] = ""
            totals_row["scenario_price"] = ""
            totals_row["days_active"] = ""
            totals_row["velocity"] = ""
            totals_row["days_to_clear"] = ""
            totals_row["batch_age_days"] = ""

            for c in total_cols:
                if c in calc_df.columns:
                    totals_row[c] = float(pd.to_numeric(calc_df[c], errors="coerce").fillna(0.0).sum())

            final_display_df = pd.concat(
                [calc_df, pd.DataFrame([totals_row])],
                ignore_index=True
            )

            st.markdown("### Scenario Worksheet Output")
            st.dataframe(final_display_df, use_container_width=True)

            try:
                _xl = _build_sku_sim_excel(
                    calc_df, D0, total_overhead_pool, days_elapsed,
                    shipment_name=selected_shipment,
                )
                st.download_button(
                    label="📥 Download Scenario as Excel (with formulas)",
                    data=_xl,
                    file_name=f"sku_sim_{selected_shipment}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as _xl_err:
                st.caption(f"Excel export unavailable: {_xl_err}")

            st.session_state["last_scenario_worksheet_df"] = calc_df.copy()

            with st.expander("📐 How SKU Simulator calculates scenario profit", expanded=False):
                st.markdown("""
**Purpose:** Test pricing and clearance scenarios using the same overhead pool from Batch Profitability — without re-running the full engine.

---
**Editable inputs (per row):**

| Input | Effect |
|---|---|
| `avg_price` | New selling price per unit for the remaining stock |
| `days_to_clear` | Revised estimate of how many days to sell the remaining units (capped at 730) |

---
**What stays fixed (anchored to the original Batch Profitability run):**
- All realized figures: `sold_qty`, `sold_revenue`, `realized_cogs`, `realized_gm`, `overhead_realized`, `net_profit_realized`
- Overhead pool: `D0 = total_overhead_pool ÷ max(days_active)` — same as original run
- `remaining_qty` and `remaining_cost_value`

---
**What recalculates from your edits:**

| Field | Formula |
|---|---|
| `proj_remaining_revenue` | `remaining_qty × new avg_price` |
| `proj_remaining_gm` | `proj_remaining_revenue − remaining_cost_value` |
| `remaining_share` | `SKU proj_remaining_revenue ÷ total new proj remaining revenue` (falls back to cost share if 0) |
| `overhead_projected` | `D0 × 0.97^(days_to_clear ÷ 60) × days_to_clear × remaining_share` |
| `Proj_remaining_profit` | `proj_remaining_gm − overhead_projected` |
| `proj_final_profit` | `net_profit_realized + Proj_remaining_profit` |

*The 0.97 decay factor means longer clearance horizons attract diminishing marginal overhead — consistent with the Batch Profitability engine.*

---
**Velocity (read-only, from Batch Profitability):**
- `velocity = sold_qty ÷ days_active` (units/day)
- If `sold_qty = 0` → velocity defaults to 0.02
- `days_to_clear` is initially set to `remaining_qty ÷ velocity` but you can override it here

---
**Numerical example:**

Original result from Batch Profitability:
- `net_profit_realized = 13,900`; `remaining_qty = 30`; `unit_cost = 500`
- `remaining_cost_value = 15,000`; `avg_price = 750`; `days_to_clear = 45`
- `D0 = 100/day`; `remaining_share = 25%`
- `proj_final_profit = 20,316` (from Batch Profitability)

**Scenario: lower price to 650:**
- `proj_remaining_revenue = 30 × 650 = 19,500`
- `proj_remaining_gm = 19,500 − 15,000 = 4,500`
- `overhead_projected = 100 × 0.97^(45/60) × 45 × 0.25 ≈ 1,084`
- `Proj_remaining_profit = 4,500 − 1,084 = 3,416`
- `proj_final_profit = 13,900 + 3,416 = 17,316` *(vs 20,316 at BDT 750)*

**Scenario: same price 750 but push clearance to 90 days:**
- `overhead_projected = 100 × 0.97^(90/60) × 90 × 0.25 ≈ 2,119`
- `Proj_remaining_profit = 7,500 − 2,119 = 5,381`
- `proj_final_profit = 13,900 + 5,381 = 19,281` *(slower clearance costs ~1,000 in overhead)*
""")
            return
